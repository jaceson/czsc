import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import akshare as ak
import time
import sys
import os

# 添加项目根目录到路径
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from czsc_daily_util import get_daily_symbols
from czsc_sqlite import get_local_stock_data

class RPSIndicator:
    """
    RPS（相对强度指标）计算器
    模拟通达信扩展数据的0-1000归一化顺序排名
    """
    
    def __init__(self, stock_codes=None):
        """
        初始化RPS计算器
        
        Args:
            stock_codes: 股票代码列表，如 ['000001', '000002', ...]
                        如果为None，则自动从 get_daily_symbols() 获取
        """
        if stock_codes is None:
            # 从 get_daily_symbols 获取所有股票
            print("正在从本地配置获取股票列表...")
            self.stock_codes = get_daily_symbols()
            print(f"已加载 {len(self.stock_codes)} 只股票")
        else:
            self.stock_codes = stock_codes
        
        self.price_data = {}  # 存储价格数据
        self.rps_data = {}    # 存储RPS数据
        
    def fetch_price_data(self, days=250, start_date='2022-01-01'):
        """
        获取股票历史价格数据（从本地数据库）
        
        Args:
            days: 获取多少天的历史数据（仅用于显示，实际从 start_date 开始）
            start_date: 起始日期，默认'2000-01-01'
        """
        print(f"开始从本地数据库获取 {len(self.stock_codes)} 只股票的历史数据...")
        
        success_count = 0
        fail_count = 0
        
        for i, code in enumerate(self.stock_codes):
            try:
                # 从本地数据库获取历史数据
                df = get_local_stock_data(code, start_date)
                
                if df is not None and not df.empty and len(df) > days:
                    # 确保有日期列
                    if 'date' in df.columns:
                        df['date'] = pd.to_datetime(df['date'])
                        df.set_index('date', inplace=True)
                    
                    # 只保留收盘价
                    if 'close' in df.columns:
                        self.price_data[code] = df['close']
                        success_count += 1
                        
                        if (i + 1) % 100 == 0:
                            print(f"✓ [{i+1}/{len(self.stock_codes)}] {code}: 加载成功，{len(df)}条数据")
                    else:
                        fail_count += 1
                        if (i + 1) % 500 == 0:
                            print(f"✗ [{i+1}/{len(self.stock_codes)}] {code}: 缺少close列")
                else:
                    fail_count += 1
                    if (i + 1) % 500 == 0:
                        print(f"✗ [{i+1}/{len(self.stock_codes)}] {code}: 数据不足或为空")
                    
            except Exception as e:
                fail_count += 1
                if (i + 1) % 500 == 0:
                    print(f"✗ [{i+1}/{len(self.stock_codes)}] {code}: 获取失败 - {e}")
        
        print(f"\n数据获取完成！")
        print(f"  成功: {success_count} 只")
        print(f"  失败: {fail_count} 只")
        print(f"  总计: {len(self.price_data)} 只股票有有效数据")
        
    def calculate_returns(self, periods=[5, 10, 20, 60, 120, 250]):
        """
        计算不同周期的收益率
        
        Args:
            periods: 周期列表（交易日），如 [5, 10, 20, 60, 120, 250]
                     对应周、半月、月、季、半年、年
        
        Returns:
            returns_data: 各周期收益率数据
        """
        returns_data = {period: {} for period in periods}
        
        for code, prices in self.price_data.items():
            for period in periods:
                if len(prices) > period:
                    # 计算收益率: (当前价格 / N日前价格 - 1) * 100
                    ret = (prices.iloc[-1] / prices.iloc[-period-1] - 1) * 100
                    returns_data[period][code] = ret
                else:
                    returns_data[period][code] = np.nan
                    
        return returns_data
    
    def calculate_normalized_rank(self, metric_dict, ascending=False):
        """
        计算0-1000归一化顺序排名
        
        Args:
            metric_dict: 指标字典，如 {'000001': 10.5, '000002': 8.3, ...}
            ascending: 排名方向
                      False: 值越大排名越高（用于涨幅，涨幅大的得高分）
                      True: 值越小排名越高（用于跌幅等）
        
        Returns:
            归一化排名得分字典，范围0-1000
        """
        # 过滤掉NaN值
        valid_items = {k: v for k, v in metric_dict.items() if not np.isnan(v)}
        
        if len(valid_items) == 0:
            return {code: 0 for code in metric_dict.keys()}
        
        # 转换为DataFrame进行排序
        df = pd.DataFrame(list(valid_items.items()), columns=['code', 'value'])
        
        # 排序
        df_sorted = df.sort_values('value', ascending=ascending)
        
        # 生成排名（1开始）
        df_sorted['rank'] = range(1, len(df_sorted) + 1)
        
        # 归一化公式: ((总数量 - 排名) / (总数量 - 1)) * 1000
        total = len(df_sorted)
        df_sorted['normalized_score'] = ((total - df_sorted['rank']) / (total - 1)) * 1000
        
        # 转换为字典
        score_dict = dict(zip(df_sorted['code'], df_sorted['normalized_score']))
        
        # 对于无效数据，得分为0
        for code in metric_dict.keys():
            if code not in score_dict:
                score_dict[code] = 0
                
        return score_dict
    
    def calculate_rps(self, periods=[5, 10, 20, 60, 120, 250]):
        """
        计算各周期的RPS值
        
        Args:
            periods: 周期列表
        
        Returns:
            rps_data: 各周期RPS数据字典
        """
        # 1. 计算各周期收益率
        returns_data = self.calculate_returns(periods)
        
        # 2. 对每个周期进行归一化排名
        rps_data = {}
        for period in periods:
            rps_scores = self.calculate_normalized_rank(returns_data[period], ascending=False)
            rps_data[period] = rps_scores
            
            # 输出统计信息
            valid_scores = [v for v in rps_scores.values() if v > 0]
            if valid_scores:
                print(f"RPS{period}: 有效数据 {len(valid_scores)} 只，"
                      f"最高 {max(valid_scores):.1f}，"
                      f"最低 {min(valid_scores):.1f}，"
                      f"平均 {np.mean(valid_scores):.1f}")
            else:
                print(f"RPS{period}: 无有效数据")
                
        return rps_data
    
    def to_dataframe(self, rps_data):
        """
        将RPS数据转换为DataFrame格式
        
        Args:
            rps_data: calculate_rps返回的数据
        
        Returns:
            DataFrame，列为股票代码，行为各周期RPS值
        """
        df = pd.DataFrame(rps_data)
        df.index.name = 'period'
        df.columns.name = 'stock_code'
        
        # 转置，使行表示股票，列表示周期
        df_t = df.T
        df_t.columns = [f'RPS{col}' for col in df_t.columns]
        
        return df_t
    
    def get_top_stocks(self, rps_data, period=20, top_n=20):
        """
        获取指定周期RPS排名靠前的股票
        
        Args:
            rps_data: calculate_rps返回的数据
            period: 周期，如20
            top_n: 返回前N只股票
        
        Returns:
            排名靠前的股票列表
        """
        if period not in rps_data:
            print(f"周期 {period} 不存在")
            return []
        
        scores = rps_data[period]
        sorted_stocks = sorted(scores.items(), key=lambda x: x[1], reverse=True)
        
        return sorted_stocks[:top_n]
    
    def save_to_excel(self, df, filename='rps_data.xlsx'):
        """
        保存RPS数据到Excel文件
        """
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='RPS数据')
            
            # 添加条件格式：高亮RPS>90的单元格
            workbook = writer.book
            worksheet = writer.sheets['RPS数据']
            
            # 创建条件格式
            from openpyxl.styles import PatternFill
            fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            
            # 对每列应用条件格式
            for col_idx, col in enumerate(df.columns, start=2):
                for row_idx in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if cell.value and cell.value > 90:
                        cell.fill = fill
        
        print(f"数据已保存到 {filename}")
    
    def quick_rps(self, periods=[20, 60, 120], start_date='2000-01-01'):
        """
        快速计算常用周期的RPS（20日、60日、120日）
        
        Args:
            periods: RPS周期列表
            start_date: 历史数据起始日期
        """
        print("="*60)
        print("开始计算RPS指标...")
        print("="*60)
        
        # 获取价格数据（需要250天用于计算120日RPS）
        self.fetch_price_data(days=250, start_date=start_date)
        
        if len(self.price_data) == 0:
            print("没有获取到任何股票数据，请检查本地数据库")
            return None
        
        # 计算RPS
        rps_data = self.calculate_rps(periods)
        
        # 转换为DataFrame
        df = self.to_dataframe(rps_data)
        
        # 添加综合评分（各周期RPS加权平均）
        weights = {20: 0.4, 60: 0.35, 120: 0.25}
        df['综合评分'] = 0
        for period, weight in weights.items():
            if period in df.columns:
                df['综合评分'] += df[f'RPS{period}'] * weight
        
        # 按综合评分排序
        df_sorted = df.sort_values('综合评分', ascending=False)
        
        # 输出前20名
        print("\n" + "="*60)
        print("RPS综合评分前20名股票")
        print("="*60)
        print(df_sorted.head(20).to_string())
        
        return df_sorted


# ========== 使用示例 ==========

def main():
    """
    使用示例：计算A股市场主要股票的RPS指标
    """
    
    # 方式1: 自动从 get_daily_symbols() 获取所有股票（推荐）
    print("方式1: 从本地配置获取所有股票列表")
    rps_calculator = RPSIndicator()
    
    # 方式2: 手动指定股票列表（测试用）
    # test_stocks = [
    #     '000001', '000002', '000004', '000006', '000008',
    #     '000009', '000012', '000016', '000021', '000025'
    # ]
    # rps_calculator = RPSIndicator(stock_codes=test_stocks)
    
    if len(rps_calculator.stock_codes) == 0:
        print("未加载到任何股票，请检查配置")
        return
    
    # 快速计算RPS（20日、60日、120日）
    # 从本地数据库读取历史数据，start_date 为起始日期
    result_df = rps_calculator.quick_rps(
        periods=[20, 60, 120],
        start_date='2000-01-01'  # 从2000年开始的数据
    )
    
    if result_df is not None:
        # 保存结果
        rps_calculator.save_to_excel(result_df, 'rps_analysis.xlsx')
        
        # 获取RPS20排名前5的股票
        print("\n" + "="*60)
        print("RPS20排名前5的股票")
        print("="*60)
        top_stocks = rps_calculator.get_top_stocks(rps_calculator.rps_data, period=20, top_n=5)
        for code, score in top_stocks:
            print(f"{code}: RPS20 = {score:.2f}")


# ========== 高级功能：从本地数据计算 ==========

class RPSFromLocalData(RPSIndicator):
    """
    从本地Excel/CSV文件读取数据计算RPS
    """
    
    def load_from_csv(self, filepath, date_col='date', price_col='close'):
        """
        从CSV文件加载价格数据
        
        Args:
            filepath: CSV文件路径
            date_col: 日期列名
            price_col: 价格列名
        """
        df = pd.read_csv(filepath)
        df[date_col] = pd.to_datetime(df[date_col])
        
        # 按股票代码分组
        for code in df['code'].unique():
            stock_df = df[df['code'] == code].set_index(date_col)
            self.price_data[code] = stock_df[price_col]
            
        print(f"从 {filepath} 加载了 {len(self.price_data)} 只股票数据")
    
    def load_from_database(self, conn, table_name):
        """
        从数据库加载价格数据
        """
        query = f"SELECT code, date, close FROM {table_name}"
        df = pd.read_sql(query, conn)
        df['date'] = pd.to_datetime(df['date'])
        
        for code in df['code'].unique():
            stock_df = df[df['code'] == code].set_index('date')
            self.price_data[code] = stock_df['close']
            
        print(f"从数据库加载了 {len(self.price_data)} 只股票数据")


if __name__ == "__main__":
    main()